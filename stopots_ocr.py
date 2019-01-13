#!/usr/bin/env python
# -*- coding: utf-8 -*-
#-------------------------------------------------------------------------------
# Name:        stopots OCR cheat
# Purpose:
# This script provides a cheat for playing the online game stopots (Brasilian
# kids game known as STOP or ADEDANHA
# Once executed, the script will grab the computer screen, try to detect the
# selected letter and the words required by the game, and will use a dictionary
# to present player with some word suggestions
#
# Author:      Alex Porto
#
# Created:     13/01/2019
# Copyright:   (c) alex 2019
# Licence:     gpl 3
#
# Dependencies:
#   - Tesseract OCR binaries (Download the Windows installer)
#   - pytesseract: python binding for tesseract
#   - PIL: python image handling lib
#   - openpyxl: python lib for reading Excel files

#-------------------------------------------------------------------------------


#-------------------------------------------------------------------------------
# User options. Configure these settings so the script will adapt to your computer

# Rectangle area on the screen where the words and the selected letter appear
# rectangle is defined by (x, y, width, height)
WORDS_AREA = (499, 344, 624, 407)
LETTER_AREA = (774, 213, 78, 76)

# if DEBUG is True, the script will read snapshot.png from disk instead of
# grabbing it from Windows screen
DEBUG = True

#-------------------------------------------------------------------------------

#-------------------------------------------------------------------------------
# Do not change anything below this line, unless you know what you're doing
#-------------------------------------------------------------------------------

#-------------------------------------------------------------------------------
# Code definitions:
#   - request: A class of word requested by the game for the current round
#       (changes every round) (E.g.: carro, CEP, fruta, etc)
#   - letter: The single character string selected by the game for this round
#       (also changes every round)
#-------------------------------------------------------------------------------

import os
import sys
import pytesseract as ocr
from PIL import Image, ImageOps, ImageGrab
from openpyxl import load_workbook
from random import randint
from unicodedata import normalize

base_path = os.path.dirname(__file__)

def remove_accents(txt, codif='utf-8'):
  return normalize('NFKD', txt.decode(codif)).encode('ASCII', 'ignore')

def load_dictionary():
    words = {}
    original_requests = {}
    excel_file = os.path.join(base_path, 'dictionary.xlsx')
    wb = load_workbook(excel_file)
    sheet = wb.get_sheet_by_name('Sheet1')
    col = 0
    while True:
        col += 1
        request = sheet.cell(1, col).value
        if request is None:
            break
        request = request.encode('utf-8').strip()
        request = remove_accents(request).upper()
        request = unicode(request, 'utf-8')
        original_request = request
        request = request.replace(' ', '') # Remove spaces to avoid OCR problems
        original_requests[request] = original_request
        words[request] = {}
        line = 1
        while True:
            line += 1
            word = sheet.cell(line, col).value
            if word is None:
                break
            word = word.encode('utf-8').strip()
            word = remove_accents(word).upper()
            word = unicode(word, 'utf-8')
            try:
                words[request][word[0]].append(word)
            except:
                words[request][word[0]] = [word]
    return words, original_requests

def decode_image():
    # TODO: Improve this hard coded location of tesseract.exe.
    ocr.pytesseract.tesseract_cmd = r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"
    if DEBUG:
        # if debugging the code, load the screenshot from file
        img = Image.open(os.path.join(base_path, 'screenshot.png'))
    else:
        # normal use of the script: Grab a screen of the game.
        # Please note that PIL ImageGrab has some limitations on multiple
        # monitors, so you shall keep your browser open at the main monitor
        # if you have two or more monitors
        img = ImageGrab.grab()
        # save a copy of the image for further debug analysis
        save_path = os.path.join(base_path, 'screenshot.png')
        img.save(save_path)

    # Stopops image has a lot of colors, so we first onvert it to black and
    # white for better OCR performance
    thresh = 150
    fn = lambda x : 255 if x > thresh else 0
    words_area = (WORDS_AREA[0], WORDS_AREA[1], WORDS_AREA[0] + WORDS_AREA[2], WORDS_AREA[1] + WORDS_AREA[3])
    letter_area = (LETTER_AREA[0], LETTER_AREA[1], LETTER_AREA[0] + LETTER_AREA[2], LETTER_AREA[1] + LETTER_AREA[3])
    words_img = img.crop(words_area)
    words_img = ImageOps.invert(words_img)
    words_img = words_img.convert('L').point(fn, mode='1')

    letter_img = img.crop(letter_area)
    letter_img = ImageOps.invert(letter_img)
    letter_img = letter_img.convert('L').point(fn, mode='1')

    save_path = os.path.join(base_path, 'letter.png')
    letter_img.save(save_path)

    # Decode a phrase containing all the request requested by the game
    phrase = ocr.image_to_string(words_img, lang='por')
    phrase = phrase.encode('utf-8')
    phrase = remove_accents(phrase).upper()
    phrase = phrase.replace('\n', ' ')
    phrase = phrase.replace('\r', ' ')
    phrase = phrase.replace('  ', ' ')


    # Decode the single letter selected by the game
    letter = ocr.image_to_string(letter_img, lang='por', config='-psm 10').upper()
    letter = str(letter)

    return phrase, letter

def main():
    words, original_requests = load_dictionary()
    phrase, letter = decode_image()

    if letter is None or letter == '':
        print "Letra nao detectada!"
        return

    print "--------------------------------------------------------------------------"
    print "Letra:", letter
    print(phrase)
    print "--------------------------------------------------------------------------"

    # Use phrase without any spaces for comparison, because OCR may insert erroneous
    # spaces inbetween the words
    joined_phrase = phrase.replace(' ', '')

    # Get all requests found at the OCRed string that also exist in
    # the word dictionary, then sort them according the order they appear on the
    # OCRed string
    found_requests = [c for c in words.keys() if joined_phrase.find(c) >= 0]
    indices = [joined_phrase.index(c) for c in found_requests]
    found_requests = [x for _,x in sorted(zip(indices, found_requests))]

    for request in found_requests:
        try:
            word_options = words[request][letter]
            word = word_options[randint(0,len(word_options)-1)]
        except:
            word = '< sem sugestao >'
        # original_requests is a dict containg the request name but
        # that still keeps the spaces between the words
        # spaces were removed to avoid OCD errors
        name = original_requests[request]
        print name + '\t: ' +  word

if __name__ == '__main__':
    main()







